#!/usr/bin/env python3
"""
Currency Conversion and Rounding Report Generator

This script analyzes ERP API integration logs to extract currency conversion data
and generates comprehensive reports for the finance team to review.

Key Features:
- Parses log files for currency conversion operations
- Extracts exchange rates, amounts, and rounding details
- Validates calculations and identifies discrepancies
- Generates Excel reports with multiple sheets for analysis
"""

import re
import logging
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import json

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class CurrencyConversionAnalyzer:
    def __init__(self, log_file_path):
        self.log_file_path = Path(log_file_path)
        self.conversion_data = []
        self.currency_transformations = []
        self.voucher_summaries = {}
        
    def parse_log_file(self):
        """Parse the ERP integration log file to extract currency conversion data."""
        logger.info(f"Parsing log file: {self.log_file_path}")
        
        with open(self.log_file_path, 'r', encoding='utf-8') as file:
            content = file.read()
            
        # Extract currency conversion operations
        self._extract_currency_conversions(content)
        
        # Extract currency transformations
        self._extract_currency_transformations(content)
        
        # Build voucher summaries
        self._build_voucher_summaries()
        
        logger.info(f"Extracted {len(self.conversion_data)} currency conversions")
        logger.info(f"Extracted {len(self.currency_transformations)} currency transformations")
        
    def _extract_currency_conversions(self, content):
        """Extract currency conversion operations from log content."""
        # Pattern for currency conversion with exchange rate
        conversion_pattern = r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d{3}).*?Converted (\d+\.?\d*) (\w+) to (\d+\.?\d*) (\w+) \(rate: (\d+\.?\d*)\)'
        
        # Pattern for raw conversion details
        raw_conversion_pattern = r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d{3}).*?Raw conversion: (\d+\.?\d*), After Decimal rounding: (\d+\.?\d*)'
        
        # Pattern for voucher context
        voucher_pattern = r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d{3}).*?Processing individual entry - Voucher: ([\w-]+)'
        
        conversions = re.findall(conversion_pattern, content)
        raw_conversions = re.findall(raw_conversion_pattern, content)
        vouchers = re.findall(voucher_pattern, content)
        
        # Build conversion data with context
        current_voucher = None
        voucher_index = 0
        
        for i, conversion in enumerate(conversions):
            timestamp, original_amount, from_currency, converted_amount, to_currency, exchange_rate = conversion
            
            # Find corresponding voucher
            while voucher_index < len(vouchers) and vouchers[voucher_index][0] <= timestamp:
                current_voucher = vouchers[voucher_index][1]
                voucher_index += 1
            
            # Find corresponding raw conversion
            raw_amount = None
            rounded_amount = None
            if i < len(raw_conversions):
                raw_amount = raw_conversions[i][1]
                rounded_amount = raw_conversions[i][2]
            
            conversion_record = {
                'timestamp': timestamp,
                'voucher_number': current_voucher,
                'original_amount': Decimal(original_amount),
                'from_currency': from_currency,
                'converted_amount': Decimal(converted_amount),
                'to_currency': to_currency,
                'exchange_rate': Decimal(exchange_rate),
                'raw_conversion': Decimal(raw_amount) if raw_amount else None,
                'rounded_amount': Decimal(rounded_amount) if rounded_amount else None,
                'rounding_difference': None
            }
            
            # Calculate rounding difference
            if conversion_record['raw_conversion'] and conversion_record['rounded_amount']:
                conversion_record['rounding_difference'] = (
                    conversion_record['rounded_amount'] - conversion_record['raw_conversion']
                )
            
            self.conversion_data.append(conversion_record)
    
    def _extract_currency_transformations(self, content):
        """Extract currency code transformations from log content."""
        # Pattern for currency transformations
        transform_pattern = r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d{3}).*?Transforming currency code for company (\w+): (\w*) -> \'?([^\']*?)\'?$'
        
        # Pattern for R- prefix operations
        r_prefix_pattern = r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d{3}).*?(Normalized currency code by removing R- prefix|Adding R- prefix to \w+): ([^-]*?) -> \'?([^\']*?)\'?$'
        
        transforms = re.findall(transform_pattern, content, re.MULTILINE)
        r_operations = re.findall(r_prefix_pattern, content, re.MULTILINE)
        
        for transform in transforms:
            timestamp, company, original, transformed = transform
            self.currency_transformations.append({
                'timestamp': timestamp,
                'operation_type': 'company_transform',
                'company': company,
                'original_currency': original,
                'transformed_currency': transformed,
                'description': f'Company {company} transformation'
            })
        
        for r_op in r_operations:
            timestamp, operation, original, transformed = r_op
            self.currency_transformations.append({
                'timestamp': timestamp,
                'operation_type': 'r_prefix',
                'company': None,
                'original_currency': original,
                'transformed_currency': transformed,
                'description': operation
            })
    
    
    def _build_voucher_summaries(self):
        """Build summary data for each voucher."""
        for conversion in self.conversion_data:
            voucher = conversion['voucher_number']
            if voucher not in self.voucher_summaries:
                self.voucher_summaries[voucher] = {
                    'voucher_number': voucher,
                    'conversions': [],
                    'total_original_amount': Decimal('0'),
                    'total_converted_amount': Decimal('0'),
                    'total_rounding_difference': Decimal('0'),
                    'currencies_involved': set(),
                    'exchange_rates_used': set()
                }
            
            summary = self.voucher_summaries[voucher]
            summary['conversions'].append(conversion)
            summary['total_original_amount'] += conversion['original_amount']
            summary['total_converted_amount'] += conversion['converted_amount']
            
            if conversion['rounding_difference']:
                summary['total_rounding_difference'] += conversion['rounding_difference']
            
            summary['currencies_involved'].add(f"{conversion['from_currency']}→{conversion['to_currency']}")
            summary['exchange_rates_used'].add(f"{conversion['from_currency']}/{conversion['to_currency']}: {conversion['exchange_rate']}")
    
    
    def generate_excel_report(self, output_path):
        """Generate comprehensive Excel report."""
        logger.info(f"Generating Excel report: {output_path}")
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self._create_summary_sheet(wb)
        self._create_detail_sheet(wb)
        self._create_transformation_sheet(wb)
        self._create_exchange_rates_sheet(wb)
        
        # Save workbook
        wb.save(output_path)
        logger.info(f"Excel report saved: {output_path}")
    
    def _create_summary_sheet(self, wb):
        """Create summary sheet with voucher-level aggregations."""
        ws = wb.create_sheet("Summary", 0)
        
        # Headers
        headers = [
            'Voucher Number', 'Total Conversions', 'Total Original Amount',
            'Total Converted Amount', 'Total Rounding Difference',
            'Currencies Involved', 'Exchange Rates Used', 'Status'
        ]
        
        ws.append(headers)
        
        # Data
        for voucher, summary in self.voucher_summaries.items():
            ws.append([
                summary['voucher_number'],
                len(summary['conversions']),
                float(summary['total_original_amount']),
                float(summary['total_converted_amount']),
                float(summary['total_rounding_difference']),
                '; '.join(summary['currencies_involved']),
                '; '.join(summary['exchange_rates_used']),
                'Reviewed' if abs(summary['total_rounding_difference']) < 0.01 else 'Needs Review'
            ])
        
        self._format_sheet(ws, headers)
    
    def _create_detail_sheet(self, wb):
        """Create detailed conversion sheet."""
        ws = wb.create_sheet("Conversion Details")
        
        # Headers
        headers = [
            'Timestamp', 'Voucher Number', 'From Currency', 'Original Amount',
            'To Currency', 'Exchange Rate', 'Raw Conversion', 'Rounded Amount',
            'Final Converted', 'Rounding Difference', 'Precision Used'
        ]
        
        ws.append(headers)
        
        # Data
        for conversion in self.conversion_data:
            ws.append([
                conversion['timestamp'],
                conversion['voucher_number'],
                conversion['from_currency'],
                float(conversion['original_amount']),
                conversion['to_currency'],
                float(conversion['exchange_rate']),
                float(conversion['raw_conversion']) if conversion['raw_conversion'] else '',
                float(conversion['rounded_amount']) if conversion['rounded_amount'] else '',
                float(conversion['converted_amount']),
                float(conversion['rounding_difference']) if conversion['rounding_difference'] else '',
                '2 decimal places'
            ])
        
        self._format_sheet(ws, headers)
    
    
    def _create_transformation_sheet(self, wb):
        """Create currency transformation sheet."""
        ws = wb.create_sheet("Currency Transformations")
        
        # Headers
        headers = [
            'Timestamp', 'Operation Type', 'Company', 'Original Currency',
            'Transformed Currency', 'Description'
        ]
        
        ws.append(headers)
        
        # Data
        for transform in self.currency_transformations:
            ws.append([
                transform['timestamp'],
                transform['operation_type'],
                transform['company'] or '',
                transform['original_currency'],
                transform['transformed_currency'],
                transform['description']
            ])
        
        self._format_sheet(ws, headers)
    
    def _create_exchange_rates_sheet(self, wb):
        """Create exchange rates summary sheet."""
        ws = wb.create_sheet("Exchange Rates")
        
        # Headers
        headers = [
            'Currency Pair', 'Exchange Rate', 'First Used', 'Last Used',
            'Usage Count', 'Total Amount Converted'
        ]
        
        ws.append(headers)
        
        # Aggregate exchange rate data
        rate_data = {}
        for conversion in self.conversion_data:
            pair = f"{conversion['from_currency']}/{conversion['to_currency']}"
            rate = conversion['exchange_rate']
            
            if pair not in rate_data:
                rate_data[pair] = {
                    'rate': rate,
                    'first_used': conversion['timestamp'],
                    'last_used': conversion['timestamp'],
                    'count': 0,
                    'total_amount': Decimal('0')
                }
            
            rate_data[pair]['count'] += 1
            rate_data[pair]['total_amount'] += conversion['original_amount']
            rate_data[pair]['last_used'] = max(rate_data[pair]['last_used'], conversion['timestamp'])
        
        # Data
        for pair, data in rate_data.items():
            ws.append([
                pair,
                float(data['rate']),
                data['first_used'],
                data['last_used'],
                data['count'],
                float(data['total_amount'])
            ])
        
        self._format_sheet(ws, headers)
    
    def _format_sheet(self, ws, headers):
        """Apply formatting to worksheet."""
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
    
    def generate_json_summary(self, output_path):
        """Generate JSON summary for programmatic access."""
        summary_data = {
            'report_generated': datetime.now().isoformat(),
            'log_file': str(self.log_file_path),
            'total_conversions': len(self.conversion_data),
            'total_transformations': len(self.currency_transformations),
            'voucher_count': len(self.voucher_summaries),
            'currency_pairs': list(set([
                f"{conv['from_currency']}/{conv['to_currency']}" 
                for conv in self.conversion_data
            ])),
            'rounding_summary': self._get_rounding_summary()
        }
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(summary_data, f, indent=2, default=str)
        
        logger.info(f"JSON summary saved: {output_path}")
    
    
    def _get_rounding_summary(self):
        """Get rounding impact summary."""
        total_rounding = sum(
            abs(conv['rounding_difference']) 
            for conv in self.conversion_data 
            if conv['rounding_difference']
        )
        
        max_rounding = max(
            (abs(conv['rounding_difference']) for conv in self.conversion_data if conv['rounding_difference']),
            default=0
        )
        
        return {
            'total_rounding_impact': float(total_rounding),
            'max_single_rounding': float(max_rounding),
            'conversions_with_rounding': sum(
                1 for conv in self.conversion_data if conv['rounding_difference']
            )
        }


def main():
    """Main function to run the currency conversion analysis."""
    import argparse
    
    parser = argparse.ArgumentParser(description='Generate currency conversion and rounding report')
    parser.add_argument('log_file', help='Path to ERP API integration log file')
    parser.add_argument('--output-dir', default='.', help='Output directory for reports')
    parser.add_argument('--excel', action='store_true', help='Generate Excel report')
    parser.add_argument('--json', action='store_true', help='Generate JSON summary')
    
    args = parser.parse_args()
    
    # Initialize analyzer
    analyzer = CurrencyConversionAnalyzer(args.log_file)
    
    # Parse log file
    analyzer.parse_log_file()
    
    # Generate timestamp for output files
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # Generate reports
    if args.excel or not (args.excel or args.json):
        excel_path = Path(args.output_dir) / f'currency_conversion_report_{timestamp}.xlsx'
        analyzer.generate_excel_report(excel_path)
    
    if args.json or not (args.excel or args.json):
        json_path = Path(args.output_dir) / f'currency_conversion_summary_{timestamp}.json'
        analyzer.generate_json_summary(json_path)
    
    # Print summary
    print(f"\n=== Currency Conversion Analysis Summary ===")
    print(f"Log file analyzed: {args.log_file}")
    print(f"Total conversions found: {len(analyzer.conversion_data)}")
    print(f"Total currency transformations: {len(analyzer.currency_transformations)}")
    print(f"Vouchers processed: {len(analyzer.voucher_summaries)}")
    
    
    rounding_summary = analyzer._get_rounding_summary()
    print(f"Total rounding impact: {rounding_summary['total_rounding_impact']}")
    print(f"Max single rounding difference: {rounding_summary['max_single_rounding']}")


if __name__ == '__main__':
    main()
